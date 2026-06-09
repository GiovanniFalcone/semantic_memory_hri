import pandas as pd

import csv
import os
import time
import json
import datetime
import math

ids = []
mod = []
turns = []
time_play = []
avg_time_until_match = []
wrong_moves = []
board_chanded_times = []
token_geo = []
token_math = []
pairs_resolved_from_geo_robot = []
pairs_resolved_from_math_robot = []
number_of_curiosities_from_geo_robot = []
number_of_curiosities_from_math_robot = []


def get_turn_number(df):
    """
    get number of turns necessary to end a game for each user
    """
    
    tmp = df.loc[df['game_ended'] == True]      
    turns.append(tmp['turn_number'].iloc[0])

def get_time(df):
    # avg time to find a pair
    tmp = df.loc[df['match'] == True]
    tmp = tmp['time_until_match'].tolist()
    avg = (str(datetime.timedelta(seconds=math.trunc(sum(map(lambda f: int(f[0])*60 + int(f[1]), map(lambda f: f.split(':'), tmp)))/len(tmp)))))
    avg_time_until_match.append(avg)
    # time to finish the game
    mysum = datetime.timedelta()
    for i in tmp:
        (m, s) = i.split(':')
        d = datetime.timedelta(minutes=int(m), seconds=int(s))
        mysum += d
    time_play.append(str(mysum))

def get_number_of_wrong_cards_provided(df):
    """
    get number of wrong cards provided by the robot during the game
    """
    tmp = df.loc[df['wrong_card'] == True]      
    wrong_moves.append(len(tmp))

def get_board_changed_times(df):
    """
    get number of times the board changed during the game
    """
    tmp = df.loc[df['board_changed'] == True]      
    board_chanded_times.append(len(tmp))

def get_token_geo(df):
    """
    get number of times the robot David (geography) played during the game
    """
    tmp = df.loc[df['subject'] == 'geography']      
    token_geo.append(round(len(tmp)))

def get_token_math(df):
    """
    get number of times the robot David (math) played during the game
    """
    tmp = df.loc[df['subject'] == 'math']      
    token_math.append(round(len(tmp)))

def get_pairs_resolved_from_geo_robot(df):
    """
    get number of pairs resolved by the robot David (geography) during the game
    """
    tmp = df.loc[(df['turn_token'] == 'geography') & (df['match'] == True)]      
    pairs_resolved_from_geo_robot.append(len(tmp))

def get_pairs_resolved_from_math_robot(df):
    """
    get number of pairs resolved by the robot Michael (math) during the game
    """
    tmp = df.loc[(df['turn_token'] == 'math') & (df['match'] == True)]      
    pairs_resolved_from_math_robot.append(len(tmp))

def get_number_of_curiosities_from_geo_robot(df):
    """
    get number of curiosities provided by the robot David (geography) during the game
    """
    tmp = df.loc[(df['robot_speech'] == 'geography')]      
    number_of_curiosities_from_geo_robot.append(len(tmp))

def get_number_of_curiosities_from_math_robot(df):
    """
    get number of curiosities provided by the robot Michael (math) during the game
    """
    tmp = df.loc[(df['robot_speech'] == 'math')]      
    number_of_curiosities_from_math_robot.append(len(tmp))

# get all csv 
path =r"..\data\\"
df_append = pd.DataFrame()
for root, dirs, files in sorted(os.walk(path)):
    for file in files:
        if(file.endswith(".csv")):
            print(os.path.join(root,file))
            # one data frame for file
            df = pd.read_csv(os.path.join(root,file), sep=';')
            print(df)

            # get id and mod
            ids.append(df['id_player'].iloc[0])
            mod.append(df['experiment_condition'].iloc[0])

            # get turn number
            get_turn_number(df)

            # get time
            get_time(df)

            # get wrong cards
            get_number_of_wrong_cards_provided(df)

            # get board changed times
            get_board_changed_times(df)

            # get token geo
            get_token_geo(df)

            # get token math
            get_token_math(df)

            # get pairs resolved from geo robot
            get_pairs_resolved_from_geo_robot(df)

            # get pairs resolved from math robot
            get_pairs_resolved_from_math_robot(df)

            # get number of curiosities from geo robot
            get_number_of_curiosities_from_geo_robot(df)

            # get number of curiosities from math robot
            get_number_of_curiosities_from_math_robot(df)


csv_struct = {
    "id": ids, 
    "experiment_condition": mod, "turns": turns, 
    "time to finish": time_play,                                        # include anche i robot
    "average time to find a pair": avg_time_until_match,                # include anche i robot
    "wrong moves": wrong_moves,
    "board changed times": board_chanded_times,
    "token geo": token_geo,                                             # conta tutte le volte che l'umano ha cliccato il bottone (il robot potrebbe anche aver giocato 0 turni)
    "token math": token_math,
    "pairs resolved from geo robot": pairs_resolved_from_geo_robot,     # conta anche quelle che ha risolto casualmente (le carte non erano note -> clicca random -> match)
    "pairs resolved from math robot": pairs_resolved_from_math_robot,
    "number of curiosities from geo robot": number_of_curiosities_from_geo_robot,
    "number of curiosities from math robot": number_of_curiosities_from_math_robot
}

# write on excel the results
keys = csv_struct.keys()

# create dataframe
df_final = pd.DataFrame(csv_struct)
# order dataframe by id
df_final['id'] = df_final['id'].astype(int)
df_final.sort_values(by='id', inplace=True)
# handle time
df_final['time to finish'] = pd.to_timedelta(df_final['time to finish'])
df_final['time to finish'] = df_final['time to finish'].astype(str)
df_final['time to finish'] = df_final['time to finish'].apply(lambda x: str(x).split()[2]) # remove 0 days


# calcola media e dev std

# conversione per il tempo da stringa a timedelta
df_final['time to finish'] = pd.to_timedelta(df_final['time to finish'])
df_final['average time to find a pair'] = pd.to_timedelta(df_final['average time to find a pair'])
# poi a secondi
df_final['time to finish_sec'] = df_final['time to finish'].dt.total_seconds()
df_final['avg_find_pair_sec'] = df_final['average time to find a pair'].dt.total_seconds()

# seleziona solo le colonne numeriche ed esclude id (per non fare la media)
numeric_cols = df_final.select_dtypes(include=['number']).columns.drop('id', errors='ignore')

# avg e std
df_media = df_final.groupby('experiment_condition')[numeric_cols].mean()
df_dev_std = df_final.groupby('experiment_condition')[numeric_cols].std() 

# crea un nuovo dataframe
df_new = pd.DataFrame(index=df_media.index)
for col in numeric_cols:
    df_new[f"{col}_avg"] = df_media[col]
    df_new[f"{col}_std"] = df_dev_std[col]

# ordina le colonne
colonne_ordinate = []
for col in numeric_cols:
    colonne_ordinate.extend([f"{col}_avg", f"{col}_std"])
df_ordinato = df_new[colonne_ordinate]

# sala su un'altra scheda dello stesso foglio
with pd.ExcelWriter("results.xlsx", engine='openpyxl') as writer:
    df_final.to_excel(writer, sheet_name="RAW", index=False, float_format="%.2f")
    
    df_ordinato.to_excel(writer, sheet_name="STATS", index=True, float_format="%.2f")

# auto adjust column width
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("results.xlsx")
centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        for cell in col:
            cell.alignment = centered
wb.save("results.xlsx")