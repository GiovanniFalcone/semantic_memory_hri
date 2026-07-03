import pandas as pd

def process_data_PSI(type_filter, name):
    df = pd.read_csv(f"Trust_{name}.csv")
    # debug
    # print("Colonne selezionate:")
    # print(df.iloc[:, 6:10])
    # print(df.iloc[:, 10:14])
    # print(df.iloc[:, 14:18])
    # solo le righe della modalità desiderata
    df = df[df["RISERVATA AL RICERCATORE\nSeleziona la condizione"] == type_filter]
    # seleziona l'id originale delle righe filtrate
    id_partecipanti = df["id"]
    # solo le colonne PSI
    df = df.iloc[:, 6:18]
    # print(f"\nDataframe filtrato per la modalità {type_filter}:")
    # print(df)
    # reverse SOC
    df["Questo robot: [non ha capacità sociali]"] = 6 - df["Questo robot: [non ha capacità sociali]"]
    # prendo le colonne SOC, HLP, TRU
    colonne_da_sommare = [(0, 4), (4, 8), (8, 12)]
    # somma le diverse scale
    for col in colonne_da_sommare:
        df[f'somma_{col[0]+1}_{col[1]}'] = df.iloc[:, col[0]:col[1]].sum(axis=1)
    # prendo solo le PSI
    colonne_nuove = df.filter(regex='somma_')
    colonne_nuove = colonne_nuove.div(4)
    letter = "D" if name == "David" else "M"
    # rinomina le colonne
    colonne_nuove.rename(columns={'somma_1_4': f'SOC_{letter}', 'somma_5_8': f'HLP_{letter}', 'somma_9_12': f'TRU_{letter}'}, inplace=True)
    colonne_nuove['robot_type'] = type_filter

    # aggiunge l'id
    colonne_nuove['id'] = id_partecipanti.values

    # print(f"\nColonne finali per la modalità {type_filter}:")
    print(colonne_nuove, "\n")

    return colonne_nuove

def process_data_trust(type_filter, name):
    df = pd.read_csv(f"Trust_{name}.csv")
    # debug
    # print("Colonne selezionate:")
    # print(df.iloc[:, 18:23])
    # print(df.iloc[:, 23:28])
    # print(df.iloc[:, 28:33])
    # print(df.iloc[:, 33:38])
    df = df[df["RISERVATA AL RICERCATORE\nSeleziona la condizione"] == type_filter]
    # seleziona l'id originale delle righe filtrate
    id_partecipanti = df["id"]
    df = df.iloc[:, 18:38]
    colonne_da_sommare = [(0, 5), (5, 10), (10, 15), (15, 20)]
    for col in colonne_da_sommare:
        df[f'somma_{col[0]+1}_{col[1]}'] = df.iloc[:, col[0]:col[1]].sum(axis=1)
    colonne_nuove = df.filter(regex='somma_')
    colonne_nuove = colonne_nuove.div(5)
    letter = "D" if name == "David" else "M"
    colonne_nuove.rename(columns={'somma_1_5': f'PR_{letter}', 'somma_6_10': f'PTC_{letter}', 'somma_11_15': f'PU_{letter}', 'somma_16_20': f'F_{letter}'}, inplace=True)
    colonne_nuove['robot_type'] = type_filter

    # aggiunge l'id
    colonne_nuove['id'] = id_partecipanti.values

    # print(f"\nColonne finali per la modalità {type_filter}:")
    print(colonne_nuove, "\n")

    return colonne_nuove

# per ogni modalità apre il file, somma le varie scale e fa la media
cc_D = process_data_PSI('CC', 'David')
cs0_D = process_data_PSI('CS0', 'David')
cs1_D = process_data_PSI('CS1', 'David')
ci_D = process_data_PSI('CI', 'David')
ii_D = process_data_PSI('II', 'David')

cc_M = process_data_PSI('CC', 'Michael')
cs0_M = process_data_PSI('CS0', 'Michael')
cs1_M = process_data_PSI('CS1', 'Michael')
ci_M = process_data_PSI('CI', 'Michael')
ii_M = process_data_PSI('II', 'Michael')

# concatena i dati
data_David = pd.concat([cc_D, cs0_D, cs1_D, ci_D, ii_D], ignore_index=False)
data_Michael = pd.concat([cc_M, cs0_M, cs1_M, ci_M, ii_M], ignore_index=False)

# per evitare il prodotto cartesiano, causa concat
data_final = pd.merge(data_David, data_Michael, on=['id', 'robot_type'])
# data_final.sort_values(by='id', inplace=True)

# ordina le colonne
columns_order = ['id', 'robot_type', 'SOC_D', 'SOC_M', 'HLP_D', 'HLP_M', 'TRU_D', 'TRU_M']
data_final = data_final[columns_order]

print(data_David, "\n")
print(data_Michael, "\n")
print(data_final, "\n")

# idem per trust
cc_D_trust = process_data_trust('CC', 'David')
cs0_D_trust = process_data_trust('CS0', 'David')
cs1_D_trust = process_data_trust('CS1', 'David')
ci_D_trust = process_data_trust('CI', 'David')
ii_D_trust = process_data_trust('II', 'David')

cc_M_trust = process_data_trust('CC', 'Michael')
cs0_M_trust = process_data_trust('CS0', 'Michael')
cs1_M_trust = process_data_trust('CS1', 'Michael')
ci_M_trust = process_data_trust('CI', 'Michael')
ii_M_trust = process_data_trust('II', 'Michael')

trust_David = pd.concat([cc_D_trust, cs0_D_trust, cs1_D_trust, ci_D_trust, ii_D_trust], ignore_index=True)
trust_Michael = pd.concat([cc_M_trust, cs0_M_trust, cs1_M_trust, ci_M_trust, ii_M_trust], ignore_index=True)

# merge dei due dataframe, usando id e robot_type come chiave
trust_final = pd.merge(trust_David, trust_Michael, on=['id', 'robot_type'])
# riordina le colonne
columns_order_trust = ['id', 'robot_type', 'PR_D', 'PR_M', 'PTC_D', 'PTC_M', 'PU_D', 'PU_M', 'F_D', 'F_M']
trust_final = trust_final[columns_order_trust]
print(trust_final, "\n")

# dataframe finale con tutte le scale PSI e Trust
data_final = pd.merge(data_final, trust_final, on=['id', 'robot_type'])
print(data_final, "\n")

# salva su excel
with pd.ExcelWriter('Questionnaire_results.xlsx') as writer:  
    data_final.to_excel(writer, sheet_name='PSI_scales', index=False)

# auto adjust column width
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("Questionnaire_results.xlsx")
centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        for cell in col:
            cell.alignment = centered
wb.save("Questionnaire_results.xlsx")